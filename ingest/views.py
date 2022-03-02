from calendar import c
from xmlrpc.client import Boolean
from django.conf import settings
from django.contrib import messages, auth
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.files.storage import FileSystemStorage
from django.shortcuts import render, redirect
from django.urls import reverse_lazy, reverse
from django.views.generic import DetailView
from django.views.generic.edit import UpdateView, DeleteView
from django.core.cache import cache
from django.core.exceptions import ObjectDoesNotExist
from django.core.mail import send_mail
from django.http import HttpResponse, JsonResponse, Http404

from django_filters.views import FilterView
from django_tables2.views import SingleTableMixin
from django_tables2 import RequestConfig
import pyexcel as pe
from openpyxl import load_workbook
import xlrd
import re
from celery.result import AsyncResult

from . import tasks
from .field_list import required_metadata, contributor_metadata, funder_metadata, publication_metadata, instrument_metadata, dataset_metadata, specimen_metadata, image_metadata, datastate_metadata
from .filters import CollectionFilter
from .forms import CollectionForm, ImageMetadataForm, DescriptiveMetadataForm, UploadForm, collection_send
from .models import UUID, Collection, ImageMetadata, DescriptiveMetadata, Project, ProjectPeople, People, Project, EventsLog, Contributor, Funder, Publication, Instrument, Dataset, Specimen, Image, DataState, Sheet
from .tables import CollectionTable, ImageMetadataTable, DescriptiveMetadataTable, CollectionRequestTable
import uuid
import datetime
import json
from datetime import datetime

def logout(request):
    messages.success(request, "You've successfully logged out")
    # XXX: this view should be separated from the the ingestion views and
    # placed with other authentication views to allow us to reuse the
    # authentication views with other apps (e.g. data exploration portal).
    auth.logout(request)
    # Send the user back to the login page when they log out.
    # XXX: we might want to use django's messaging system to inform them that
    # they've successfully logged out.
    return redirect('login')

def signup(request):
    """ Info about signing up for a new account. """
    # XXX: this view should be separated from the the ingestion views and
    # placed with other authentication views to allow us to reuse the
    # authentication views with other apps (e.g. data exploration portal).
    return render(request, 'ingest/signup.html')

@login_required
def index(request):
    """ The main/home page. """
    current_user = request.user
    try:
        people = People.objects.get(auth_user_id_id = current_user.id)
        project_person = ProjectPeople.objects.filter(people_id = people.id).all()
        if people.is_bil_admin:
            return render(request, 'ingest/bil_index.html', {'people':people})
        for attribute in project_person: 
            if attribute.is_pi:
                return render(request, 'ingest/pi_index.html', {'project_person': attribute})
    except Exception as e:
        print(e)
    return render(request, 'ingest/index.html')


@login_required
def pi_index(request):
    current_user = request.user
    try:
        people = People.objects.get(auth_user_id_id = current_user.id)
        project_person = ProjectPeople.objects.filter(people_id = people.id).all()
        
        for attribute in project_person:
            if project_person.is_pi:
                return render(request, 'ingest/pi_index.html', {'project_person': attribute})
    except Exception as e:
        print(e)
    return render(request, 'ingest/index.html')

# this function presents all users for changing of PI and PO
@login_required
def modify_user(request, pk):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False
    person = People.objects.get(auth_user_id_id = pk)
    all_project_people = ProjectPeople.objects.filter(people_id_id=person.id).all()   
    for project_person in all_project_people:
        try:
            their_project = Project.objects.get(id=project_person.project_id_id)
        except Project.DoesNotExist:
            their_project = None
            return render(request, 'ingest/no_projects.html', {'pi':pi})
        project_person.their_project = their_project
    return render(request, 'ingest/modify_user.html', {'all_project_people':all_project_people, 'person':person})    

# this function presents all users and gives a bil admin the option to add or remove bil admin privs to said users
@login_required
def modify_biladmin_privs(request, pk):
    # use pk to find the user in the people table
    person = People.objects.get(auth_user_id_id = pk)
    return render(request, 'ingest/modify_biladmin_privs.html', {'person':person})

# this function does the actual changing of bil admin privs
@login_required
def change_bil_admin_privs(request):
    content = json.loads(request.body)
    items = []
    for item in content:
        items.append(item['is_bil_admin'])
        is_bil_admin = item['is_bil_admin']
        person_id = item['person_id']
        
        person = People.objects.get(id=person_id)
        person.is_bil_admin=is_bil_admin
        person.save()
    return HttpResponse(json.dumps({'url': reverse('ingest:index')}))

# this function lists all the users so a pi can assign people to their project
@login_required
def list_all_users(request):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False
    allusers = User.objects.all()
    return render(request, 'ingest/list_all_users.html', {'allusers':allusers, 'pi':pi})

# this function does the actual changing of is_pi or is_po of users
@login_required
def userModify(request):
    content = json.loads(request.body)
    items = []
    for item in content:
        items.append(item['is_pi'])
        items.append(item['is_po'])
        items.append(item['auth_id'])
        items.append(item['project_id'])
        is_pi = item['is_pi']
        is_po = item['is_po']
        auth_id = item['auth_id']
        project_id = item['project_id']
        
        project_person = ProjectPeople.objects.get(id=project_id)
        project_person.is_pi=is_pi
        project_person.is_po=is_po
        project_person.save()
        
    return HttpResponse(json.dumps({'url': reverse('ingest:index')}))

@login_required
def manageProjects(request):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id, is_pi = True).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False
    
    allprojects=[]
    for row in project_person:
        project_id = row.project_id_id
        project =  Project.objects.get(id=project_id)
        allprojects.append(project)     
      
    return render(request, 'ingest/manage_projects.html', {'allprojects':allprojects, 'pi':pi})

# this functions allows pi to see all the collections
@login_required
def manageCollections(request):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False
    # gathers all the collections associated with the PI, linked on pi_index.html
    collections = []
    allprojects = ProjectPeople.objects.filter(people_id_id=people.id, is_pi=True).all()
    for proj in allprojects:
        project = Project.objects.get(id = proj.project_id_id)
        collection = Collection.objects.filter(project_id=project.id).all()
        collections.extend(collection)
    return render(request, 'ingest/manage_collections.html', {'pi':pi, 'collections':collections})

# add a new project
@login_required
def project_form(request):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False
    return render(request, 'ingest/project_form.html', {'pi':pi})

# takes the data from project_form
@login_required
def create_project(request):
    new_project = json.loads(request.body)
    items = []
    for item in new_project:
        items.append(item['funded_by'])
        items.append(item['is_biccn'])
        items.append(item['name'])
        
        funded_by = item['funded_by']
        is_biccn = item['is_biccn']
        name = item['name']
        
        # write project to the project table   
        project = Project(funded_by=funded_by, is_biccn=is_biccn, name=name)
        project.save()
        
        # create a project_people row for this pi so they can view project on pi dashboard
        project_id_id = project.id
        current_user = request.user
        person = People.objects.get(auth_user_id_id=current_user)
        
        project_person = ProjectPeople(project_id_id=project_id_id, people_id_id=person.id, is_pi=True, is_po=False, doi_role='creator')
        project_person.save()
    messages.success(request, 'Project Created!')    
    return HttpResponse(json.dumps({'url': reverse('ingest:manage_projects')}))


@login_required
def add_project_user(request, pk):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False
    all_users = User.objects.all()
    project = Project.objects.get(id=pk) 
    return render(request, 'ingest/add_project_user.html', {'all_users':all_users, 'project':project, 'pi':pi})

# adds person to a project
@login_required
def write_user_to_project_people(request):
    content = json.loads(request.body)
    items = []
    for item in content:
        items.append(item['user_id'])
        items.append(item['project_id'])
        user_id = item['user_id']
        project_id = item['project_id']

        project = Project.objects.get(id=project_id) 
        person = People.objects.get(auth_user_id_id=user_id)
        project_person = ProjectPeople(project_id_id=project.id, people_id_id=person.id, is_pi=False, is_po=False, doi_role='')
        
        try:
            check =  ProjectPeople.objects.get(project_id_id=project.id, people_id_id=person.id)
            user = User.objects.get(id=user_id)
        except:
            project_person.save()
    messages.success(request, 'User(s) Added!')
    return HttpResponse(json.dumps({'url': reverse('ingest:manage_projects')}))

# presents all people on the projects of the pi who is logged in
@login_required
def people_of_pi(request):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False
    
    pi = People.objects.get(auth_user_id_id=current_user.id)
    # filters the project_people table down to the rows where it's the pi's people_id_id AND is_pi=true
    pi_projects = ProjectPeople.objects.filter(people_id_id=pi.id, is_pi=True).all()
    for proj in pi_projects:
        proj.related_project_people = ProjectPeople.objects.filter(project_id=proj.project_id_id).all()
    return render(request, 'ingest/people_of_pi.html', {'pi_projects':pi_projects, 'pi':pi})


@login_required
def view_project_people(request, pk):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False
    try:
        project = Project.objects.get(id=pk)
        # get all of the project people rows with the project_id matching the project.id
        projectpeople = ProjectPeople.objects.filter(project_id_id=pk).all()
        # get all of the people who are in those projectpeople rows
        allpeople = []
        for row in projectpeople:
            person_id = row.people_id_id
            person = People.objects.get(id=person_id)
            allpeople.append(person)
        return render(request, 'ingest/view_project_people.html', { 'project':project, 'allpeople':allpeople })
    except ProjectPeople.DoesNotExist:
        return render(request, 'ingest/no_people.html')
    return render(request, 'ingest/view_project_people.html', {'allpeople':allpeople, 'project':project, 'pi':pi})

# fallback for when a project has no collections associated with it
@login_required
def no_collection(request, pk):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False
    project = Project.objects.get(id=pk)
    return(request, 'ingest/no_collection.html',  {'project':project, 'pi':pi})

# fallback for when a project has no people assigned to it
@login_required
def no_people(request, pk):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False
    project = Project.objects.get(id=pk)
    return(request, 'ingest/no_people.html', {'project':project, 'pi':pi})

# view all the collections of a project
@login_required
def view_project_collections(request, pk):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False
    try:
        project = Project.objects.get(id=pk)
        project_collections = Collection.objects.filter(project_id=project.id).all()
        
        for collection in project_collections:
            user_id = collection.user_id
            owner = User.objects.get(id=user_id)
            try:
                event = EventsLog.objects.filter(collection_id_id=collection.id).latest('event_type')
            except EventsLog.DoesNotExist:
                event = None
            collection.event = event
            collection.owner = owner
       
    except Collection.DoesNotExist:
        return render(request, 'ingest/no_collection.html')  
    return render(request, 'ingest/view_project_collections.html', {'project':project, 'project_collections':project_collections, 'pi':pi})

@login_required
def descriptive_metadata_list(request):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False
    """ A list of all the metadata the user has created. """
    # The user is trying to delete the selected metadata
    for key in request.POST:
        messages.success(request, key) 
        messages.success(request, request.POST[key])
        print (request.POST[key])      
    if request.method == "POST":
        pks = request.POST.getlist("selection")
        # Get all of the checked metadata (except LOCKED metadata)
        selected_objects = DescriptiveMetadata.objects.filter(
            pk__in=pks, locked=False)
        selected_objects.delete()
        messages.success(request, 'Descriptive Metadata successfully deleted')
        return redirect('ingest:descriptive_metadata_list')
    # This is the GET (just show the user their list of metadata)
    else:
        # XXX: This exclude is likely redundant, becaue there's already the
        # same exclude in the class itself. Need to test though.
        table = DescriptiveMetadataTable(
            DescriptiveMetadata.objects.filter(user=request.user), exclude=['user'])
        RequestConfig(request).configure(table)
        descriptive_metadata = DescriptiveMetadata.objects.filter(user=request.user)
    
        datasets_list = []
        collections = Collection.objects.filter(user=request.user)
        for c in collections:
            sheets = Sheet.objects.filter(collection_id=c.id).all()
            for s in sheets:
                datasets = Dataset.objects.filter(sheet_id=s.id)
                for d in datasets:
                    datasets_list.append(d)

        return render(
            request,
            'ingest/descriptive_metadata_list.html',
            {'table': table, 'descriptive_metadata': descriptive_metadata, 'pi':pi, 'datasets_list':datasets_list})

def new_metadata_detail(request, pk):
    contributors = Contributor.objects.filter(sheet_id=pk).all()
    funders = Funder.objects.filter(sheet_id=pk).all 
    publications = Publication.objects.filter(sheet_id=pk).all()
    instruments = Instrument.objects.filter(sheet_id=pk).all()
    datasets = Dataset.objects.filter(sheet_id=pk).all()
    specimens = Specimen.objects.filter(sheet_id=pk).all()
    images = Image.objects.filter(sheet_id=pk).all()
    
    return render(request, 'ingest/new_metadata_detail.html', {'contributors':contributors, 'funders':funders, 'publications':publications, 'instruments':instruments, 'datasets':datasets, 'specimens':specimens, 'images':images})

class DescriptiveMetadataDetail(LoginRequiredMixin, DetailView):
    """ A detailed view of a single piece of metadata. """
    model = DescriptiveMetadata
    template_name = 'ingest/descriptive_metadata_detail.html'
    context_object_name = 'descriptive_metadata'

@login_required
def collection_send(request):
    content = json.loads(request.body)
    items = []
    user_name = request.user
    for item in content:
        items.append(item['bil_uuid'])
        coll = Collection.objects.get(bil_uuid = item['bil_uuid'])
        coll_id = Collection.objects.get(id = coll.id)
        person = People.objects.get(name = user_name)
        person_id = person.id
        time = datetime.now()
        event = EventsLog(collection_id = coll_id, people_id_id = person.id, project_id_id = coll.project_id, notes = '', timestamp = time, event_type = 'request_validation')
        event.save()
    if request.method == "POST":
        subject = '[BIL Validations] New Validation Request'
        sender = 'ltuite96@psc.edu'
        message = F'The following collections have been requested to be validated {items} by {user_name}@psc.edu'
        recipient = ['ltuite96@psc.edu']
        
        send_mail(
        subject,
        message,
        sender,
        recipient
             )
    messages.success(request, 'Request succesfully sent')
    return HttpResponse(json.dumps({'url': reverse('ingest:index')}))

@login_required
def collection_create(request):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False
    """ Create a collection. """
    # We cache the staging area location, so that we can show it in the GET and
    # later use it during creation (POST)
    if cache.get('host_and_path'):
        host_and_path = cache.get('host_and_path')
        data_path = cache.get('data_path')
        bil_uuid = cache.get('bil_uuid')
        bil_user = cache.get('bil_user')
    else:
        top_level_dir = settings.STAGING_AREA_ROOT
        #shortens uuid
        uuidhex = (uuid.uuid4()).hex
        str1 = uuidhex[0:16]
        str2 = uuidhex[16:]
        str3 = "%x" % (int(str1,16)^int(str2,16))
        bil_uuid = str3.zfill(16)
        #this should make uuid unique or just redirect to home page if collision.
        uu = UUID(useduuid=bil_uuid)
        try:
           uu.save()
        except:
           return redirect('ingest:index') 
        #data_path = "{}/bil_data/{}/{:02d}/{}".format(
        #    top_level_dir,
        #    datetime.datetime.now().year,
        #    datetime.datetime.now().month,
        #    str(uuid.uuid4()))
        data_path = "{}/{}/{}".format(
            top_level_dir,
            request.user,
            bil_uuid)
        host_and_path = "{}@{}:{}".format(
            request.user, settings.IMG_DATA_HOST, data_path)
        bil_user = "{}".format(request.user)
        cache.set('host_and_path', host_and_path, 30)
        cache.set('data_path', data_path, 30)
        cache.set('bil_uuid', bil_uuid, 30)
        cache.set('bil_user', bil_user, 30)

    if request.method == "POST":
        # We need to pass in request here, so we can use it to get the user
        form = CollectionForm(request.POST, request=request)
        if form.is_valid():
            # remotely create the directory on some host using fabric and
            # celery
            # note: you should authenticate with ssh keys, not passwords
            if not settings.FAKE_STORAGE_AREA:
                tasks.create_data_path.delay(data_path,bil_user)
            post = form.save(commit=False)
            #post.data_path = host_and_path
            post.data_path = data_path
            post.bil_uuid = bil_uuid
            post.bil_user = bil_user
            post.save()

            time = datetime.now()
            coll_id = Collection.objects.get(id = post.id)
            #coll_id = Collection.objects.filter(bil_uuid = bil_uuid).values_list('id', flat=True)
            proj_id = coll_id.project_id
            
            event = EventsLog(collection_id = coll_id, people_id_id = people.id, project_id_id = proj_id, notes = '', timestamp = time, event_type = 'collection_created')
            event.save()
            cache.delete('host_and_path')
            cache.delete('data_path')
            cache.delete('bil_uuid')
            cache.delete('bil_user')
            messages.success(request, 'Collection successfully created!! Please proceed with metadata upload')
            return redirect('ingest:descriptive_metadata_upload')
    else:
        form = CollectionForm()

    project_list = []
    person = People.objects.get(auth_user_id=request.user.id)
    projects = ProjectPeople.objects.filter(people_id=people.id).all()
    for proj in projects:
        project = Project.objects.get(id=proj.project_id_id)  
        project_list.append(project)
  
    collections = Collection.objects.all()
    funder_list = [
        "1-U01-H114812-01",
        "1-U01-MH114819-01",
        "1-U01-MH114824-01",
        "1-U01-MH114825-01",
        "1-U01-MH114829-01",
        "1-U19-MH114821-01",
        "1-U19-MH114830-01",
        "1-U19-MH114831-01",
        "1-U24-MH114827-01",
        "1R24MH114788-01",
        "1R24MH114793-01",
    ]

    return render(
        request,
        'ingest/collection_create.html',
        {'form': form,
         'projects': projects,
         'project_list': project_list, 
         'collections': collections,
         'funder_list': funder_list,
         'host_and_path': host_and_path,
         'pi': pi})

class SubmitValidateCollectionList(LoginRequiredMixin, SingleTableMixin, FilterView):
    """ A list of all a user's collections. """

    table_class = CollectionTable
    model = Collection
    template_name = 'ingest/submit_validate_collection_list.html'
    filterset_class = CollectionFilter

    def get_queryset(self, **kwargs):
        return Collection.objects.filter(user=self.request.user)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['collections'] = Collection.objects.filter(user=self.request.user)
        return context

    def get_filterset_kwargs(self, filterset_class):
        """ Sets the default collection filter status. """
        kwargs = super().get_filterset_kwargs(filterset_class)
        if kwargs["data"] is None:
            kwargs["data"] = {"submit_status": "NOT_SUBMITTED"}
        return kwargs

class SubmitRequestCollectionList(LoginRequiredMixin, SingleTableMixin, FilterView):
    def IsPi(request):
        current_user = request.user
        people = People.objects.get(auth_user_id_id = current_user.id)
        project_person = ProjectPeople.objects.filter(people_id = people.id).all()
        for attribute in project_person:
            if attribute.is_pi:
                pi = True
            else:
                pi = False
        return render(request, {'pi':pi})
    """ A list of all a user's collections. """

    table_class = CollectionRequestTable
    model = Collection
    template_name = 'ingest/submit_request_collection_list.html'
    filterset_class = CollectionFilter

    def get_queryset(self, **kwargs):
        return Collection.objects.filter(user=self.request.user)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['collections'] = Collection.objects.filter(user=self.request.user)
        return context

    def get_filterset_kwargs(self, filterset_class):
        """ Sets the default collection filter status. """
        kwargs = super().get_filterset_kwargs(filterset_class)
        if kwargs["data"] is None:
            kwargs["data"] = {"submit_status": "NOT_SUBMITTED"}
        return kwargs
    success_url = reverse_lazy('ingest:collection_list')

class CollectionList(LoginRequiredMixin, SingleTableMixin, FilterView):
    """ A list of all a user's collections. """

    table_class = CollectionTable
    model = Collection
    template_name = 'ingest/collection_list.html'
    filterset_class = CollectionFilter

    def get_queryset(self, **kwargs):
        return Collection.objects.filter(user=self.request.user)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['collections'] = Collection.objects.filter(user=self.request.user)
        return context

    def get_filterset_kwargs(self, filterset_class):
        """ Sets the default collection filter status. """
        kwargs = super().get_filterset_kwargs(filterset_class)
        if kwargs["data"] is None:
            kwargs["data"] = {"submit_status": "NOT_SUBMITTED"}
        return kwargs
         
@login_required
def collection_data_path(request, pk):
    """ Info about the staging area for a user's collection. """

    collection = Collection.objects.get(id=pk)
    host_and_path = collection.data_path
    data_path = host_and_path.split(":")[1]

    return render(
        request,
        'ingest/collection_data_path.html',
        {'collection': collection,
         'host_and_path': host_and_path,
         'data_path': data_path})

@login_required
def collection_validation_results(request, pk):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False
    """ Where a user can see the results of validation. """
    collection = Collection.objects.get(id=pk)

    collection.validation_status = "NOT_VALIDATED"
    dir_size = ""
    outvalidfile = ""
    filecontents = ""
    invalid_metadata_directories = []
    if collection.celery_task_id_validation:
        result = AsyncResult(collection.celery_task_id_validation)
        state = result.state
        if state == 'SUCCESS':
            analysis_results = result.get()
            if analysis_results['valid']:
                collection.validation_status = "SUCCESS"
            else:
                collection.validation_status = "FAILED"
                invalid_metadata_directories = analysis_results["invalid_metadata_directories"]
            dir_size = analysis_results['dir_size']
            outvalidfile = analysis_results['output']
            #Open the log file and read the contents
            f=open(outvalidfile, "r")
            if f.mode == 'r':
               filecontents=f.read()
            f.close()
        else:
            collection.validation_status = "PENDING"

    return render(
        request,
        'ingest/collection_validation_results.html',
        {'collection': collection,
         'outfile': outvalidfile,
         'output': filecontents,
         'dir_size': dir_size,
         'invalid_metadata_directories': invalid_metadata_directories,
         'pi': pi})

@login_required
def collection_submission_results(request, pk):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False
    """ Where a user can see the results of submission. """
    collection = Collection.objects.get(id=pk)

    collection.submission_status = "NOT_SUBMITTED"
    dir_size = ""
    outvalidfile = ""
    filecontents = ""
    invalid_metadata_directories = []
    if collection.celery_task_id_submission:
        result = AsyncResult(collection.celery_task_id_submission)
        state = result.state
        if state == 'SUCCESS':
            analysis_results = result.get()
            if analysis_results['valid']:
                collection.submission_status = "SUCCESS"
            else:
                collection.submission_status = "FAILED"
                invalid_metadata_directories = analysis_results["invalid_metadata_directories"]
            dir_size = analysis_results['dir_size']
            outvalidfile = analysis_results['output']
            #Open the log file and read the contents
            f=open(outvalidfile, "r")
            if f.mode == 'r':
               filecontents=f.read()
            f.close()
        else:
            collection.submission_status = "PENDING"

    return render(
        request,
        'ingest/collection_submission_results.html',
        {'collection': collection,
         'outfile': outvalidfile,
         'output': filecontents,
         'dir_size': dir_size,
         'invalid_metadata_directories': invalid_metadata_directories,
         'pi': pi})

@login_required
def collection_detail(request, pk):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False
    """ View, edit, delete, create a particular collection. """
    # If user tries to go to a page using a collection primary key that doesn't
    # exist, give a 404
    try:
        collection = Collection.objects.get(id=pk)
    except ObjectDoesNotExist:
        raise Http404
    # the metadata associated with this collection
    #image_metadata_queryset = collection.imagemetadata_set.all()
    descriptive_metadata_queryset = collection.descriptivemetadata_set.all()
    # this is what is triggered if the user hits "Upload to this Collection"
    if request.method == 'POST' and 'spreadsheet_file' in request.FILES:
        spreadsheet_file = request.FILES['spreadsheet_file']
        upload_spreadsheet(spreadsheet_file, collection, request)
        return redirect('ingest:collection_detail', pk=pk)
    # this is what is triggered if the user hits "Submit collection"
    elif request.method == 'POST' and 'validate_collection' in request.POST:
        #---trying to model validate_collection this after submit_collection
        # lock everything (collection and associated image metadata) during
        # submission and validation. if successful, keep it locked
        collection.locked = False
        metadata_dirs = []
        for im in descriptive_metadata_queryset:
        #for im in image_metadata_queryset:
            im.locked = True
            im.save()
            metadata_dirs.append(im.r24_directory)
            #metadata_dirs.append(im.directory)
        # This is just a very simple test, which will be replaced with some
        # real validation and analysis in the future
        if not settings.FAKE_STORAGE_AREA:
            data_path = collection.data_path.__str__()
            task = tasks.run_validate.delay(data_path, metadata_dirs)
            collection.celery_task_id_validation = task.task_id
        collection.save()
        return redirect('ingest:collection_detail', pk=pk)
    elif request.method == 'POST' and 'submit_collection' in request.POST:
        # lock everything (collection and associated image metadata) during
        # submission and validation. if successful, keep it locked
        collection.locked = True
        metadata_dirs = []
        for im in descriptive_metadata_queryset:
        #for im in image_metadata_queryset:
            im.locked = True
            im.save()
            metadata_dirs.append(im.r24_directory)
        #    metadata_dirs.append(im.directory)
        # This is just a very simple test, which will be replaced with some
        # real validation and analysis in the future
        if not settings.FAKE_STORAGE_AREA:
            data_path = collection.data_path.__str__()
            task = tasks.run_analysis.delay(data_path, metadata_dirs)
            collection.celery_task_id_submission = task.task_id
        collection.save()
        return redirect('ingest:collection_detail', pk=pk)

    # check submission status
    if collection.celery_task_id_submission:
        result = AsyncResult(collection.celery_task_id_submission)
        state = result.state
        if state == 'SUCCESS':
            analysis_results = result.get()
            if analysis_results['valid']:
                collection.submission_status = "SUCCESS"
            else:
                collection.submission_status = "FAILED"
                # need to unlock, so user can fix problem
                collection.locked = False
                #for im in image_metadata_queryset:
                for im in descriptive_metadata_queryset:
                    im.locked = False
                    im.save()
        else:
            collection.submission_status = "PENDING"
    collection.save()

   # check validation status
    if collection.celery_task_id_validation:
        result = AsyncResult(collection.celery_task_id_validation)
        state = result.state
        if state == 'SUCCESS':
            analysis_results = result.get()
            if analysis_results['valid']:
                collection.validation_status = "SUCCESS"
            else:
                collection.validation_status = "FAILED"
                # need to unlock, so user can fix problem
                collection.locked = False
                #for im in image_metadata_queryset:
                for im in descriptive_metadata_queryset:
                    im.locked = False
                    im.save()
        else:
            collection.validation_status = "PENDING"
    collection.save()
    table = DescriptiveMetadataTable(
        DescriptiveMetadata.objects.filter(user=request.user, collection=collection))
    return render(
        request,
        'ingest/collection_detail.html',
        {'table': table,
         'collection': collection,
         'descriptive_metadata_queryset': descriptive_metadata_queryset,
         'pi': pi})

class CollectionUpdate(LoginRequiredMixin, UpdateView):
    """ Edit an existing collection ."""
    model = Collection
    fields = [
        'name', 'description', 'organization_name', 'lab_name',
        'project_funder', 'project_funder_id'
    ]
    
    def IsPi(request):
        current_user = request.user
        people = People.objects.get(auth_user_id_id = current_user.id)
        project_person = ProjectPeople.objects.filter(people_id = people.id).all()
        for attribute in project_person:
            if attribute.is_pi:
                pi = True
            else:
                pi = False
        return render(request, {'pi':pi})
    template_name = 'ingest/collection_update.html'
    success_url = reverse_lazy('ingest:collection_list')

@login_required
def collection_delete(request, pk):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False
    """ Delete a collection. """

    collection = Collection.objects.get(pk=pk)
    if request.method == 'POST':
        if collection.submission_status != "SUCCESS":
            data_path = collection.data_path.__str__()
            if not settings.FAKE_STORAGE_AREA:
            # This is what deletes the actual directory associated with the
            # staging area

                tasks.delete_data_path.delay(data_path)
            collection.delete()
            messages.success(request, 'Collection successfully deleted')
            return redirect('ingest:collection_list')
        else:
            messages.error(request, 'This collection is public, it cannot be deleted. If this is incorrect contact us at bil-support@psc.edu')
    return render(
        request, 'ingest/collection_delete.html', {'collection': collection, 'pi':pi})

def check_contributors_sheet(filename):
    errormsg=""
    workbook=xlrd.open_workbook(filename)
    sheetname = 'Contributors'
    contributors_sheet = workbook.sheet_by_name(sheetname)
    colheads=['contributorName','Creator','contributorType',
                 'nameType','nameIdentifier','nameIdentifierScheme',
                 'affiliation', 'affiliationIdentifier', 'affiliationIdentifierScheme']
    creator = ['Yes', 'No']
    contributortype = ['ProjectLeader','ResearchGroup','ContactPerson', 'DataCollector', 'DataCurator', 'ProjectLeader', 'ProjectManager', 'ProjectMember','RelatedPerson', 'Researcher', 'ResearchGroup','Other' ]
    nametype = ['Personal', 'Organizational']
    nameidentifierscheme = ['ORCID','ISNI','ROR','GRID','RRID' ]
    affiliationidentifierscheme = ['ORCID','ISNI','ROR','GRID','RRID' ]
    cellcols=['A','B','C','D','E','F','G','H','I']
    cols=contributors_sheet.row_values(2)
    for i in range(0,len(colheads)):
        if cols[i] != colheads[i]:
            errormsg = errormsg + ' Tab: "Contributors" cell heading found: "' + cols[i] + \
                       '" but expected: "' + colheads[i] + '" at cell: "' + cellcols[i] + '3". '
    if errormsg != "":
        return [ True, errormsg ]
    for i in range(6,contributors_sheet.nrows):
        cols=contributors_sheet.row_values(i)
        if cols[0] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[0] + '" value expected but not found in cell: "' + cellcols[0] + str(i+1) + '". '
        if cols[1] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[1] + '" value expected but not found in cell: "' + cellcols[1] + str(i+1) + '". '
        if cols[1] not in creator:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[1] + '" incorrect CV value found: "' + cols[1] + '" in cell "' + cellcols[1] + str(i+1) + '". '
        if cols[2] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[2] + '" value expected but not found in cell "' + cellcols[2] + str(i+1) + '". '
        if cols[2] not in contributortype:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[2] + '" incorrect CV value found: "' + cols[2] + '" in cell "' + cellcols[2] + str(i+1) + '". '
        if cols[3] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[3] + '" value expected but not found in cell "' + cellcols[3] + str(i+1) + '". '
        if cols[3] not in nametype:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[3] + '" incorrect CV value found: "' + cols[3] + '" in cell "' + cellcols[3] + str(i+1) + '". '
        if cols[3] == "Personal":
            if cols[4] == "":
                errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[4] + '" value expected but not found in cell "' + cellcols[4] + str(i+1) + '". '
            if cols[5] == "":
                errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[5] + '" value expected but not found in cell "' + cellcols[5] + str(i+1) + '". '
            if cols[5] not in nameidentifierscheme:
                errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[5] + '" incorrect CV value found: "' + cols[5] + '" in cell "' + cellcols[5] + str(i+1) + '". '
        if cols[6] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[6] + '" value expected but not found in cell "' + cellcols[6] + str(i+1) + '". '
        if cols[7] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[7] + '" value expected but not found in cell "' + cellcols[7] + str(i+1) + '". '
        if cols[8] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[8] + '" value expected but not found in cell "' + cellcols[8] + str(i+1) + '". '
        if cols[8] not in affiliationidentifierscheme:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[8] + '" Incorrect CV value found: "' + cols[8] + '" in cell "' + cellcols[8] + str(i+1) + '". '
    return errormsg

def check_funders_sheet(filename):
    errormsg=""
    workbook=xlrd.open_workbook(filename)
    sheetname = 'Funders'
    funders_sheet = workbook.sheet_by_name(sheetname)
    colheads=['funderName','fundingReferenceIdentifier','fundingReferenceIdentifierType',
                 'awardNumber','awardTitle']
    fundingReferenceIdentifierType = ['ROR', 'GRID', 'ORCID', 'ISNI']
    cellcols=['A','B','C','D','E']
    cols=funders_sheet.row_values(3)
    for i in range(0,len(colheads)):
        if cols[i] != colheads[i]:
            errormsg = errormsg + ' Tab: "Funders" cell heading found: "' + cols[i] + \
                       '" but expected: "' + colheads[i] + '" at cell: "' + cellcols[i] + '3". '
    if errormsg != "":
        return [ True, errormsg ]
    for i in range(6,funders_sheet.nrows):
        cols=funders_sheet.row_values(i)
        if cols[0] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[0] + '" value expected but not found in cell: "' + cellcols[0] + str(i+1) + '". '
        if cols[1] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[1] + '" value expected but not found in cell: "' + cellcols[1] + str(i+1) + '". '
        if cols[2] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[2] + '" value expected but not found in cell "' + cellcols[2] + str(i+1) + '". '
        if cols[2] not in fundingReferenceIdentifierType:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[2] + '" incorrect CV value found: "' + cols[2] + '" in cell "' + cellcols[2] + str(i+1) + '". '
        if cols[3] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[3] + '" value expected but not found in cell "' + cellcols[3] + str(i+1) + '". '
        if cols[4] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[4] + '" value expected but not found in cell "' + cellcols[4] + str(i+1) + '". '
    return errormsg

def check_publication_sheet(filename):
    errormsg=""
    workbook=xlrd.open_workbook(filename)
    sheetname = 'Publication'
    publication_sheet = workbook.sheet_by_name(sheetname)
    colheads=['relatedIdentifier','relatedIdentifierType','PMCID',
                 'relationType','citation']
    relatedIdentifierType = ['arcXiv', 'DOI', 'PMID', 'ISBN']
    relationType = ['IsCitedBy', 'IsDocumentedBy']
    cellcols=['A','B','C','D','E']
    cols=publication_sheet.row_values(3)
    for i in range(0,len(colheads)):
        if cols[i] != colheads[i]:
            errormsg = errormsg + ' Tab: "Publication" cell heading found: "' + cols[i] + \
                       '" but expected: "' + colheads[i] + '" at cell: "' + cellcols[i] + '3". '
    if errormsg != "":
        return [ True, errormsg ]
    for i in range(6,publication_sheet.nrows):
        cols=publication_sheet.row_values(i)
        if cols[0] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[0] + '" value expected but not found in cell: "' + cellcols[0] + str(i+1) + '". '
        if cols[1] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[1] + '" value expected but not found in cell: "' + cellcols[1] + str(i+1) + '". '
        if cols[1] not in relatedIdentifierType:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[1] + '" incorrect CV value found: "' + cols[1] + '" in cell "' + cellcols[1] + str(i+1) + '". '
        if cols[2] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[2] + '" value expected but not found in cell "' + cellcols[2] + str(i+1) + '". '
        if cols[3] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[3] + '" value expected but not found in cell "' + cellcols[3] + str(i+1) + '". '
        if cols[3] not in relationType:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[3] + '" incorrect CV value found: "' + cols[3] + '" in cell "' + cellcols[3] + str(i+1) + '". '
        if cols[4] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[4] + '" value expected but not found in cell "' + cellcols[4] + str(i+1) + '". '
    return errormsg

def check_instrument_sheet(filename):
    errormsg=""
    workbook=xlrd.open_workbook(filename)
    sheetname = 'Instrument'
    instrument_sheet = workbook.sheet_by_name(sheetname)
    colheads=['MicroscopeType','MicroscopeManufacturerAndModel','ObjectiveName',
                 'ObjectiveImmersion','ObjectiveNA', 'ObjectiveMagnification', 'DetectorType', 'DetectorModel', 'IlluminationTypes', 'IlluminationWavelength', 'DetectionWavelength', 'SampleTemperature']
    cellcols=['A','B','C','D','E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
    cols=instrument_sheet.row_values(3)
    for i in range(0,len(colheads)):
        if cols[i] != colheads[i]:
            errormsg = errormsg + ' Tab: "Instrument" cell heading found: "' + cols[i] + \
                       '" but expected: "' + colheads[i] + '" at cell: "' + cellcols[i] + '3". '
    if errormsg != "":
        return [ True, errormsg ]
    for i in range(6,instrument_sheet.nrows):
        cols=instrument_sheet.row_values(i)
        if cols[0] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[0] + '" value expected but not found in cell: "' + cellcols[0] + str(i+1) + '". '
        if cols[1] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[1] + '" value expected but not found in cell: "' + cellcols[1] + str(i+1) + '". '
        if cols[2] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[1] + '" value expected but not found in cell: "' + cellcols[1] + str(i+1) + '". '
        if cols[3] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[3] + '" value expected but not found in cell "' + cellcols[3] + str(i+1) + '". '
        if cols[4] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[4] + '" value expected but not found in cell "' + cellcols[4] + str(i+1) + '". '
        if cols[5] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[5] + '" value expected but not found in cell "' + cellcols[5] + str(i+1) + '". '
        if cols[6] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[6] + '" value expected but not found in cell "' + cellcols[6] + str(i+1) + '". '
        if cols[7] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[7] + '" value expected but not found in cell "' + cellcols[7] + str(i+1) + '". '
        if cols[8] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[8] + '" value expected but not found in cell "' + cellcols[8] + str(i+1) + '". '
        if cols[9] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[9] + '" value expected but not found in cell "' + cellcols[9] + str(i+1) + '". '
        if cols[10] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[10] + '" value expected but not found in cell "' + cellcols[10] + str(i+1) + '". '
        if cols[11] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[11] + '" value expected but not found in cell "' + cellcols[11] + str(i+1) + '". '
    return errormsg

def check_dataset_sheet(filename):
    errormsg=""
    workbook=xlrd.open_workbook(filename)
    sheetname = 'Dataset'
    dataset_sheet = workbook.sheet_by_name(sheetname)
    colheads=['BILDirectory','title','socialMedia','subject',
                 'Subjectscheme','rights', 'rightsURI', 'rightsIdentifier', 'Image', 'GeneralModality', 'Technique', 'Other', 'Abstract', 'Methods', 'TechnicalInfo']
    GeneralModality = ['cell morphology', 'connectivity', 'population imaging', 'spatial transcriptomics', 'other', 'anatomy', 'histology imaging', 'multimodal']
    Technique = ['anterograde tracing', 'retrograde transynaptic tracing', 'TRIO tracing', 'smFISH', 'DARTFISH', 'MERFISH', 'Patch-seq', 'fMOST', 'other', 'cre-dependent anterograde tracing','enhancer virus labeling', 'FISH', 'MORF genetic sparse labeling', 'mouselight', 'neuron morphology reconstruction', 'Patch-seq', 'retrograde tracing', 'retrograde transsynaptic tracing', 'seqFISH', 'STPT', 'VISor', 'confocal microscopy']
    cellcols=['A','B','C','D','E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
    cols=dataset_sheet.row_values(3)
    for i in range(0,len(colheads)):
        if cols[i] != colheads[i]:
            errormsg = errormsg + ' Tab: "Dataset" cell heading found: "' + cols[i] + \
                       '" but expected: "' + colheads[i] + '" at cell: "' + cellcols[i] + '3". '
    if errormsg != "":
        return [ True, errormsg ]
    for i in range(6,dataset_sheet.nrows):
        cols=dataset_sheet.row_values(i)
        if cols[0] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[0] + '" value expected but not found in cell: "' + cellcols[0] + str(i+1) + '". '
        if cols[1] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[1] + '" value expected but not found in cell: "' + cellcols[1] + str(i+1) + '". '
        if cols[2] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[1] + '" value expected but not found in cell: "' + cellcols[1] + str(i+1) + '". '
        if cols[3] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[3] + '" value expected but not found in cell "' + cellcols[3] + str(i+1) + '". '
        if cols[4] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[4] + '" value expected but not found in cell "' + cellcols[4] + str(i+1) + '". '
        if cols[5] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[5] + '" value expected but not found in cell "' + cellcols[5] + str(i+1) + '". '
        if cols[6] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[6] + '" value expected but not found in cell "' + cellcols[6] + str(i+1) + '". '
        if cols[7] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[7] + '" value expected but not found in cell "' + cellcols[7] + str(i+1) + '". '
        if cols[8] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[8] + '" value expected but not found in cell "' + cellcols[8] + str(i+1) + '". '
        if cols[9] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[9] + '" value expected but not found in cell "' + cellcols[9] + str(i+1) + '". '
        if cols[9] not in GeneralModality:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[9] + '" incorrect CV value found: "' + cols[9] + '" in cell "' + cellcols[9] + str(i+1) + '". '
        if cols[10] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[10] + '" value expected but not found in cell "' + cellcols[10] + str(i+1) + '". '
        if cols[10] not in Technique:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[10] + '" incorrect CV value found: "' + cols[10] + '" in cell "' + cellcols[10] + str(i+1) + '". '
        if cols[11] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[11] + '" value expected but not found in cell "' + cellcols[11] + str(i+1) + '". '
        if cols[12] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[12] + '" value expected but not found in cell "' + cellcols[12] + str(i+1) + '". '
        if cols[13] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[13] + '" value expected but not found in cell "' + cellcols[13] + str(i+1) + '". '
        if cols[14] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[14] + '" value expected but not found in cell "' + cellcols[14] + str(i+1) + '". '
    return errormsg

def check_specimen_sheet(filename):
    errormsg=""
    workbook=xlrd.open_workbook(filename)
    sheetname = 'Specimen'
    specimen_sheet = workbook.sheet_by_name(sheetname)
    colheads=['LocalID', 'Species', 'NCBITaxonomy', 'Age', 'Ageunit', 'Sex', 'Genotype', 'OrganLocalID', 'OrganName', 'SampleLocalID', 'Atlas', 'Locations']
    Sex = ['Male', 'Female', 'Unknown']
    cellcols=['A','B','C','D','E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
    cols=specimen_sheet.row_values(3)
    for i in range(0,len(colheads)):
        if cols[i] != colheads[i]:
            errormsg = errormsg + ' Tab: "Specimen" cell heading found: "' + cols[i] + \
                       '" but expected: "' + colheads[i] + '" at cell: "' + cellcols[i] + '3". '
    if errormsg != "":
        return [ True, errormsg ]
    for i in range(6,specimen_sheet.nrows):
        cols=specimen_sheet.row_values(i)
        if cols[0] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[0] + '" value expected but not found in cell: "' + cellcols[0] + str(i+1) + '". '
        if cols[1] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[1] + '" value expected but not found in cell: "' + cellcols[1] + str(i+1) + '". '
        if cols[2] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[1] + '" value expected but not found in cell: "' + cellcols[1] + str(i+1) + '". '
        if cols[3] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[3] + '" value expected but not found in cell "' + cellcols[3] + str(i+1) + '". '
        if cols[4] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[4] + '" value expected but not found in cell "' + cellcols[4] + str(i+1) + '". '
        if cols[5] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[5] + '" value expected but not found in cell "' + cellcols[5] + str(i+1) + '". '
        if cols[5] not in Sex:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[5] + '" incorrect CV value found: "' + cols[5] + '" in cell "' + cellcols[6] + str(i+1) + '". '
        if cols[6] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[6] + '" value expected but not found in cell "' + cellcols[6] + str(i+1) + '". '
        if cols[7] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[7] + '" value expected but not found in cell "' + cellcols[7] + str(i+1) + '". '
        if cols[8] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[8] + '" value expected but not found in cell "' + cellcols[8] + str(i+1) + '". '
        if cols[9] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[9] + '" value expected but not found in cell "' + cellcols[9] + str(i+1) + '". '
        if cols[10] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[10] + '" value expected but not found in cell "' + cellcols[10] + str(i+1) + '". '
        if cols[11] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[11] + '" value expected but not found in cell "' + cellcols[11] + str(i+1) + '". '
    return errormsg

def check_image_sheet(filename):
    errormsg=""
    workbook=xlrd.open_workbook(filename)
    sheetname = 'Image'
    image_sheet = workbook.sheet_by_name(sheetname)
    colheads=['xAxis','obliqueXdim1','obliqueXdim2',
                 'obliqueXdim3','yAxis', 'obliqueYdim1', 'obliqueYdim2', 'obliqueYdim3', 'zAxis', 'obliqueZdim1', 'obliqueZdim2', 'obliqueZdim3', 'landmarkName', 'landmarkX', 'landmarkY', 'landmarkZ', 'Number', 'displayColor', 'Representation', 'Flurophore', 'stepSizeX', 'stepSizeY', 'stepSizeZ', 'stepSizeT', 'Channels', 'Slices', 'z', 'Xsize', 'Ysize', 'Zsize', 'Gbytes', 'Files', 'DimensionOrder']
    ObliqueZdim3 = ['Superior', 'Inferior']
    ObliqueZdim2 = ['Anterior', 'Posterior']
    ObliqueZdim1 = ['Right', 'Left']
    zAxis = ['right-to-left', 'left-to-right', 'anterior-to-posterior', 'posterior-to-anterior', 'superior-to-inferior', 'inferior-to-superior', 'oblique']
    obliqueYdim3 = ['Superior', 'Inferior']
    obliqueYdim2 = ['Anterior', 'Posterior']
    obliqueYdim1 = ['Right', 'Left']
    yAxis = ['right-to-left', 'left-to-right', 'anterior-to-posterior', 'posterior-to-anterior', 'superior-to-inferior', 'inferior-to-superior', 'oblique']
    obliqueXdim3 = ['Superior', 'Inferior']
    obliqueXdim2 = ['Anterior', 'Posterior']
    obliqueXdim1 = ['Right', 'Left']
    xAxis = ['right-to-left', 'left-to-right', 'anterior-to-posterior', 'posterior-to-anterior', 'superior-to-inferior', 'inferior-to-superior', 'oblique']

    cellcols=['A','B','C','D','E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG']
    cols=image_sheet.row_values(3)
    for i in range(0,len(colheads)):
        if cols[i] != colheads[i]:
            errormsg = errormsg + ' Tab: "Image" cell heading found: "' + cols[i] + \
                       '" but expected: "' + colheads[i] + '" at cell: "' + cellcols[i] + '3". '
    if errormsg != "":
        return [ True, errormsg ]
    for i in range(6,image_sheet.nrows):
        cols=image_sheet.row_values(i)
        if cols[0] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname + 'Column: "' + colheads[0] + '" value expected but not found in cell: "' + cellcols[0] + str(i+1) + '". '
        if cols[0] not in xAxis:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname + 'Column: "' + colheads[0] + '" incorrect CV value found: "' + cols[0] + '" in cell "' + cellcols[0] + str(i+1) + '". '
        if cols[1] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname + 'Column: "' + colheads[1] + '" value expected but not found in cell: "' + cellcols[1] + str(i+1) + '". '
        if cols[1] not in obliqueXdim1:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[1] + '" incorrect CV value found: "' + cols[1] + '" in cell "' + cellcols[1] + str(i+1) + '". '
        if cols[2] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[1] + '" value expected but not found in cell: "' + cellcols[2] + str(i+1) + '". '
        if cols[2] not in obliqueXdim2:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[2] + '" incorrect CV value found: "' + cols[2] + '" in cell "' + cellcols[2] + str(i+1) + '". '
        if cols[3] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[3] + '" value expected but not found in cell "' + cellcols[3] + str(i+1) + '". '
        if cols[3] not in obliqueXdim3:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[3] + '" incorrect CV value found: "' + cols[3] + '" in cell "' + cellcols[3] + str(i+1) + '". '
        if cols[4] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[4] + '" value expected but not found in cell "' + cellcols[4] + str(i+1) + '". '
        if cols[4] not in yAxis:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[4] + '" incorrect CV value found: "' + cols[4] + '" in cell "' + cellcols[4] + str(i+1) + '". '
        if cols[5] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[5] + '" value expected but not found in cell "' + cellcols[5] + str(i+1) + '". '
        if cols[5] not in obliqueYdim1:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[5] + '" incorrect CV value found: "' + cols[5] + '" in cell "' + cellcols[5] + str(i+1) + '". '
        if cols[6] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[6] + '" value expected but not found in cell "' + cellcols[6] + str(i+1) + '". '
        if cols[6] not in obliqueYdim2:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[6] + '" incorrect CV value found: "' + cols[6] + '" in cell "' + cellcols[6] + str(i+1) + '". '
        if cols[7] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[7] + '" value expected but not found in cell "' + cellcols[7] + str(i+1) + '". '
        if cols[7] not in obliqueYdim3:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[7] + '" incorrect CV value found: "' + cols[7] + '" in cell "' + cellcols[7] + str(i+1) + '". '
        if cols[8] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[8] + '" value expected but not found in cell "' + cellcols[8] + str(i+1) + '". '
        if cols[8] not in zAxis:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[8] + '" incorrect CV value found: "' + cols[8] + '" in cell "' + cellcols[8] + str(i+1) + '". '
        if cols[9] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[9] + '" value expected but not found in cell "' + cellcols[9] + str(i+1) + '". '
        if cols[9] not in ObliqueZdim1:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[9] + '" incorrect CV value found: "' + cols[9] + '" in cell "' + cellcols[9] + str(i+1) + '". '
        if cols[10] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[10] + '" value expected but not found in cell "' + cellcols[10] + str(i+1) + '". '
        if cols[10] not in ObliqueZdim2:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[10] + '" incorrect CV value found: "' + cols[10] + '" in cell "' + cellcols[10] + str(i+1) + '". '
        if cols[11] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[11] + '" value expected but not found in cell "' + cellcols[11] + str(i+1) + '". '
        if cols[11] not in ObliqueZdim3:
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[11] + '" incorrect CV value found: "' + cols[11] + '" in cell "' + cellcols[11] + str(i+1) + '". '
        if cols[12] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[12] + '" value expected but not found in cell "' + cellcols[12] + str(i+1) + '". '
        if cols[13] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[13] + '" value expected but not found in cell "' + cellcols[13] + str(i+1) + '". '
        if cols[14] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[12] + '" value expected but not found in cell "' + cellcols[12] + str(i+1) + '". '
        if cols[15] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[15] + '" value expected but not found in cell "' + cellcols[15] + str(i+1) + '". '
        if cols[16] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[12] + '" value expected but not found in cell "' + cellcols[12] + str(i+1) + '". '
        if cols[17] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[12] + '" value expected but not found in cell "' + cellcols[12] + str(i+1) + '". '
        if cols[18] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[18] + '" value expected but not found in cell "' + cellcols[18] + str(i+1) + '". '
        if cols[19] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[12] + '" value expected but not found in cell "' + cellcols[12] + str(i+1) + '". '
        if cols[20] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[20] + '" value expected but not found in cell "' + cellcols[20] + str(i+1) + '". '
        if cols[21] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[21] + '" value expected but not found in cell "' + cellcols[21] + str(i+1) + '". '
        if cols[22] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[22] + '" value expected but not found in cell "' + cellcols[22] + str(i+1) + '". '
        if cols[23] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[23] + '" value expected but not found in cell "' + cellcols[23] + str(i+1) + '". '
        if cols[24] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[24] + '" value expected but not found in cell "' + cellcols[24] + str(i+1) + '". '
        if cols[25] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[25] + '" value expected but not found in cell "' + cellcols[25] + str(i+1) + '". '
        if cols[26] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[26] + '" value expected but not found in cell "' + cellcols[26] + str(i+1) + '". '
        if cols[27] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[27] + '" value expected but not found in cell "' + cellcols[27] + str(i+1) + '". '
        if cols[28] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[28] + '" value expected but not found in cell "' + cellcols[28] + str(i+1) + '". '
        if cols[29] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[29] + '" value expected but not found in cell "' + cellcols[29] + str(i+1) + '". '
        if cols[30] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[30] + '" value expected but not found in cell "' + cellcols[30] + str(i+1) + '". '
        if cols[31] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[31] + '" value expected but not found in cell "' + cellcols[31] + str(i+1) + '". '
        if cols[32] == "":
            errormsg = errormsg + 'On spreadsheet tab:' + sheetname +  'Column: "' + colheads[32] + '" value expected but not found in cell "' + cellcols[32] + str(i+1) + '". '
    return errormsg

def ingest_contributors_sheet(filename):
    fn = xlrd.open_workbook(filename)
    contributors_sheet = fn.sheet_by_name('Contributors')
    keys = [contributors_sheet.cell(2, col).value for col in range(contributors_sheet.ncols)]
    contributors = []
    for row in range(6,contributors_sheet.nrows):
        values = {keys[col]: contributors_sheet.cell(row, col).value
            for col in range(contributors_sheet.ncols)}
        contributors.append(values)
    return contributors

def ingest_funders_sheet(filename):
    fn = xlrd.open_workbook(filename)
    funders_sheet = fn.sheet_by_name('Funders')
    keys = [funders_sheet.cell(3, col).value for col in range(funders_sheet.ncols)]   
    funders = []
    for row in range(6, funders_sheet.nrows):
        values={keys[col]: funders_sheet.cell(row,col).value
            for col in range(funders_sheet.ncols)}
        funders.append(values)
    return funders

def ingest_publication_sheet(filename):
    fn = xlrd.open_workbook(filename)
    publication_sheet = fn.sheet_by_name('Publication')
    keys = [publication_sheet.cell(3, col).value for col in range(publication_sheet.ncols)]   
    publications = []
    for row in range(6, publication_sheet.nrows):
        values={keys[col]: publication_sheet.cell(row,col).value
            for col in range(publication_sheet.ncols)}
        publications.append(values)
    return publications

def ingest_instrument_sheet(filename):
    fn = xlrd.open_workbook(filename)
    instrument_sheet = fn.sheet_by_name('Instrument')
    keys = [instrument_sheet.cell(3,col).value for col in range(instrument_sheet.ncols)]
    instruments = []
    for row in range(6, instrument_sheet.nrows):
        values={keys[col]: instrument_sheet.cell(row,col).value
            for col in range(instrument_sheet.ncols)}
        instruments.append(values)
    return instruments

def ingest_dataset_sheet(filename):
    fn = xlrd.open_workbook(filename)
    dataset_sheet = fn.sheet_by_name('Dataset')
    keys = [dataset_sheet.cell(3,col).value for col in range(dataset_sheet.ncols)]
    datasets = []
    for row in range(6, dataset_sheet.nrows):
        values={keys[col]: dataset_sheet.cell(row,col).value
            for col in range(dataset_sheet.ncols)}
        datasets.append(values)
    return datasets

def ingest_specimen_sheet(filename):
    fn = xlrd.open_workbook(filename)
    specimen_sheet = fn.sheet_by_name('Specimen')
    keys = [specimen_sheet.cell(3,col).value for col in range(specimen_sheet.ncols)] 
    specimen_set = []
    for row in range(6, specimen_sheet.nrows):
        values={keys[col]: specimen_sheet.cell(row,col).value
            for col in range(specimen_sheet.ncols)}
        specimen_set.append(values)
    return specimen_set

def ingest_image_sheet(filename):
    fn = xlrd.open_workbook(filename)
    image_sheet = fn.sheet_by_name('Image')
    keys = [image_sheet.cell(3,col).value for col in range(image_sheet.ncols)]
    images = []
    for row in range(6, image_sheet.nrows):
        values={keys[col]: image_sheet.cell(row,col).value
            for col in range(image_sheet.ncols)}
        images.append(values)
    return images

def save_sheet_row(filename, collection):
    try:
        sheet = Sheet(filename=filename, date_uploaded=datetime.now(), collection_id=collection.id)
        sheet.save()
    except Exception as e:
        print(e)
    return sheet

def save_contributors_sheet(contributors, sheet):
    try:
        for c in contributors:
            contributorname = c['contributorName']
            creator = c['Creator']
            contributortype = c['contributorType']
            nametype = c['nameType']
            nameidentifier = c['nameIdentifier']
            nameidentifierscheme = c['nameIdentifierScheme']
            affiliation = c['affiliation']
            affiliationidentifier = c['affiliationIdentifier']
            affiliationidentifierscheme = c['affiliationIdentifierScheme']
            
            contributor = Contributor(contributorname=contributorname, creator=creator, contributortype=contributortype, nametype=nametype, nameidentifier=nameidentifier, nameidentifierscheme=nameidentifierscheme, affiliation=affiliation, affiliationidentifier=affiliationidentifier, affiliationidentifierscheme=affiliationidentifierscheme, sheet_id=sheet.id)
            contributor.save()
    except Exception as e:
        print(repr(e))
    return

def save_funders_sheet(funders, sheet):
    try:
        for f in funders:
            fundername = f['funderName']
            funding_reference_identifier = f['fundingReferenceIdentifier']
            funding_reference_identifier_type = f['fundingReferenceIdentifierType']
            award_number = f['awardNumber']
            award_title = f['awardTitle']
            
            funder = Funder(fundername=fundername, funding_reference_identifier=funding_reference_identifier, funding_reference_identifier_type=funding_reference_identifier_type, award_number=award_number, award_title=award_title, sheet_id=sheet.id)
            funder.save()
    except Exception as e:
        print(repr(e))
    return

def save_publication_sheet(publications, sheet):
    try:
        for p in publications:
            relatedidentifier = p['relatedIdentifier']
            relatedidentifiertype = p['relatedIdentifierType']
            pmcid = p['PMCID']
            relationtype = p['relationType']
            citation = p['citation']
            
            publication = Publication(relatedidentifier=relatedidentifier, relatedidentifiertype=relatedidentifiertype, pmcid=pmcid, relationtype=relationtype, citation=citation, sheet_id=sheet.id)
            publication.save()
    except Exception as e:
        print(repr(e))
    return

def save_instrument_sheet(instruments, sheet):
    try:
        for i in instruments:
            microscopetype = i['MicroscopeType']
            microscopemanufacturerandmodel = i['MicroscopeManufacturerAndModel']
            objectivename = i['ObjectiveName']
            objectiveimmersion = i['ObjectiveImmersion']
            objectivena = i['ObjectiveNA']
            objectivemagnification = i['ObjectiveMagnification']
            detectortype = i['DetectorType']
            detectormodel = i['DetectorModel']
            illuminationtypes = i['IlluminationTypes']
            illuminationwavelength = i['IlluminationWavelength']
            detectionwavelength = i['DetectionWavelength']
            sampletemperature = i['SampleTemperature']
            
            instrument = Instrument(microscopetype=microscopetype, microscopemanufacturerandmodel=microscopemanufacturerandmodel, objectivename=objectivename, objectiveimmersion=objectiveimmersion, objectivena=objectivena, objectivemagnification=objectivemagnification, detectortype=detectortype, detectormodel=detectormodel, illuminationtypes=illuminationtypes, illuminationwavelength=illuminationwavelength, detectionwavelength=detectionwavelength, sampletemperature=sampletemperature, sheet_id=sheet.id)
            instrument.save()
    except Exception as e:
        print(repr(e))
    return

def save_dataset_sheet(datasets, sheet):
    try:
        for d in datasets:
            bildirectory = d['BILDirectory']
            title = d['title']
            socialmedia = d['socialMedia']
            subject = d['subject']
            subjectscheme = d['Subjectscheme']
            rights = d['rights']
            rightsuri = d['rightsURI']
            rightsidentifier = d['rightsIdentifier']
            image = d['Image']
            generalmodality = d['GeneralModality']
            technique = d['Technique']
            other = d['Other']
            abstract = d['Abstract']
            methods = d['Methods']
            technicalinfo = d['TechnicalInfo']

            dataset = Dataset(bildirectory=bildirectory, title=title, socialmedia=socialmedia, subject=subject, subjectscheme=subjectscheme, rights=rights, rightsuri=rightsuri, rightsidentifier=rightsidentifier, image=image, generalmodality=generalmodality, technique=technique, other=other, abstract=abstract, methods=methods, technicalinfo=technicalinfo, sheet_id=sheet.id)
            dataset.save()
    except Exception as e:
        print(repr(e))
    return

def save_specimen_sheet(specimen_set, sheet):
    try:
        for s in specimen_set:
            localid = s['LocalID']
            species = s['Species']
            ncbitaxonomy = s['NCBITaxonomy']
            age = s['Age']
            ageunit = s['Ageunit']
            sex = s['Sex']
            genotype = s['Genotype']
            organlocalid = s['OrganLocalID']
            organname = s['OrganName']
            samplelocalid = s['SampleLocalID']
            atlas = s['Atlas']
            locations = s['Locations']

            specimen_object = Specimen(localid=localid, species=species, ncbitaxonomy=ncbitaxonomy, age=age, ageunit=ageunit, sex=sex, genotype=genotype, organlocalid=organlocalid, organname=organname, samplelocalid=samplelocalid, atlas=atlas, locations=locations, sheet_id=sheet.id)
            specimen_object.save()
    except Exception as e:
        print(repr(e))
    return

def save_image_sheet(images, sheet):
    try:
        for i in images:
            xaxis = i['xAxis']
            obliquexdim1 = i['obliqueXdim1']
            obliquexdim2 = i['obliqueXdim2']
            obliquexdim3 = i['obliqueXdim3']
            yaxis = i['yAxis']
            obliqueydim1 = i['obliqueYdim1']
            obliqueydim2 = i['obliqueYdim2']
            obliqueydim3 = i['obliqueYdim3']
            zaxis = i['zAxis']
            obliquezdim1 = i['obliqueZdim1']
            obliquezdim2 = i['obliqueZdim2']
            obliquezdim3 = i['obliqueZdim3']
            landmarkname = i['landmarkName']
            landmarkx = i['landmarkX']
            landmarky = i['landmarkY']
            landmarkz = i['landmarkY']
            number = i['Number']
            displaycolor = i['displayColor']
            representation = i['Representation']
            flurophore = i['Flurophore']
            stepsizex = i['stepSizeX']
            stepsizey = i['stepSizeY']
            stepsizez = i['stepSizeZ']
            stepsizet = i['stepSizeT']
            channels = i['Channels']
            slices = i['Slices']
            z = i['z']
            xsize = i['Xsize']
            ysize = i['Ysize']
            zsize = i['Zsize']
            gbytes = i['Gbytes']
            files = i['Files']
            dimensionorder = i['DimensionOrder']
    
            image = Image(xaxis=xaxis, obliquexdim1=obliquexdim1, obliquexdim2=obliquexdim2, obliquexdim3=obliquexdim3, yaxis=yaxis, obliqueydim1=obliqueydim1, obliqueydim2=obliqueydim2, obliqueydim3=obliqueydim3, zaxis=zaxis, obliquezdim1=obliquezdim1, obliquezdim2=obliquezdim2, obliquezdim3=obliquezdim3,landmarkname=landmarkname, landmarkx=landmarkx, landmarky=landmarky, landmarkz=landmarkz, number=number, displaycolor=displaycolor, representation=representation, flurophore=flurophore, stepsizex=stepsizex, stepsizey=stepsizey, stepsizez=stepsizez, stepsizet=stepsizet, channels=channels, slices=slices, z=z, xsize=xsize, ysize=ysize, zsize=zsize, gbytes=gbytes, files=files, dimensionorder=dimensionorder, sheet_id=sheet.id)
            image.save()
    except Exception as e:
        print(repr(e))
    return

def check_all_sheets(filename):
    errormsg = check_contributors_sheet(filename)
    if errormsg != '':
        return errormsg
    errormsg = check_funders_sheet(filename)
    if errormsg != '':
        return errormsg
    errormsg = check_publication_sheet(filename)
    if errormsg != '':
        return errormsg
    errormsg = check_instrument_sheet(filename)
    if errormsg != '':
        return errormsg
    errormsg = check_dataset_sheet(filename)
    if errormsg != '':
        return errormsg
    errormsg = check_specimen_sheet(filename)
    if errormsg != '':
        return errormsg
    errormsg = check_image_sheet(filename)
    if errormsg != '':
        return errormsg
    return errormsg

def save_all_sheets(contributors, funders, publications, instruments, datasets, specimen_set, images, filename, collection):
    saved = Boolean
    try:
        sheet = save_sheet_row(filename, collection)
        save_contributors_sheet(contributors, sheet)
        save_funders_sheet(funders, sheet)
        save_publication_sheet(publications, sheet)
        save_instrument_sheet(instruments, sheet)
        save_dataset_sheet(datasets, sheet)
        save_specimen_sheet(specimen_set, sheet)
        save_image_sheet(images, sheet)
        saved = True
        return saved
    except Exception as e:
        print(repr(e))
        saved = False
        return saved

def metadata_version_check(filename):
    version1 = False
    workbook=xlrd.open_workbook(filename)
    if workbook.sheet_by_name('README'):
        version1 = False
    else:
        version1 = True
    return version1

@login_required
def descriptive_metadata_upload(request):
    current_user = request.user
    people = People.objects.get(auth_user_id_id = current_user.id)
    print(people.id)
    project_person = ProjectPeople.objects.filter(people_id = people.id).all()
    for attribute in project_person:
        if attribute.is_pi:
            pi = True
        else:
            pi = False    
    """ Upload a spreadsheet containing image metadata information. """
    # The POST. A user has selected a file and associated collection to upload.
    if request.method == 'POST' and request.FILES['spreadsheet_file']:
        form = UploadForm(request.POST)
        if form.is_valid():
            associated_collection = form.cleaned_data['associated_collection']
            print(associated_collection)
            # for production
            #datapath=collection.data_path.replace("/lz/","/etc/")
            
            # for development on vm
            # datapath = '/home/shared_bil_dev/testetc/' 

            # for development locally
            datapath = '/Users/ecp/Desktop/bil_metadata_uploads' 
            
            spreadsheet_file = request.FILES['spreadsheet_file']

            fs = FileSystemStorage(location=datapath)
            name_with_path=datapath + '/' + spreadsheet_file.name
            filename = fs.save(name_with_path, spreadsheet_file)
            
            version1 = metadata_version_check(filename)
            if version1 == True:
                error = upload_descriptive_spreadsheet(filename, associated_collection, request)
                if error:
                    return redirect('ingest:descriptive_metadata_upload')
                else:         
                    return redirect('ingest:descriptive_metadata_list')
            elif version1 == False:
                errormsg = check_all_sheets(filename)
                if errormsg != '':
                    messages.error(request, errormsg)
                    return redirect('ingest:descriptive_metadata_upload')
                else:
                    collection = Collection.objects.get(name=associated_collection)
                    contributors = ingest_contributors_sheet(filename)
                    funders = ingest_funders_sheet(filename)
                    publications = ingest_publication_sheet(filename)
                    instruments = ingest_instrument_sheet(filename)
                    datasets = ingest_dataset_sheet(filename)
                    specimen_sets = ingest_specimen_sheet(filename)
                    images = ingest_image_sheet(filename)

                    saved = save_all_sheets(contributors, funders, publications, instruments, datasets, specimen_sets, images, filename, collection)
                    
                    if saved == True:
                        messages.success(request, 'Descriptive Metadata successfully uploaded!!')
                        return redirect('ingest:descriptive_metadata_list')
                    else:
                        messages.error(request, 'There has been an error. Please contact BIL Support')
                        return redirect('ingest:descriptive_metadata_upload')


    # This is the GET (just show the metadata upload page)
    else:
        user = request.user
        form = UploadForm()
        # Only let a user associate metadata with an unlocked collection that
        # they own
        form.fields['associated_collection'].queryset = Collection.objects.filter(
            locked=False, user=request.user)
        collections = form.fields['associated_collection'].queryset
    collections = Collection.objects.filter(locked=False, user=request.user)
    
    return render( request, 'ingest/descriptive_metadata_upload.html',{'form': form, 'pi':pi, 'collections':collections})

def upload_descriptive_spreadsheet(filename, associated_collection, request):
    """ Helper used by image_metadata_upload and collection_detail."""
    worksheet = filename.sheet_by_index(0)
    error = False
    try:
        missing = False
        badgrantnum = False
        has_escapes = False
        missing_fields = []
        missing_cells = []
        badchar = "\\"
        bad_str = []
        not_missing = []
        grantpattern = '[A-Z0-9\-][A-Z0-9\-][A-Z0-9A]{3}\-[A-Z0-9]{8}\-[A-Z0-9]{2}'
        for rowidx in range(worksheet.nrows):
            row = worksheet.row(rowidx)
            for colidx, cell in enumerate(row):
                if rowidx == 0:
                    #Bad Headers Check
                    #if cell.value == 'grant_number':
                    #    grantrow = rowidx+1
                    #    grantcol = colidx
                    #    grantnum = worksheet.cell(grantrow, grantcol).value
                    #    if bool(re.match(grantpattern, grantnum)) != True:
                    #        badgrantnum=True
                    # this will check specifically the headers of the document for missing info.
                    if cell.value not in required_metadata:
                        missing = True
                        missingcol = colidx+1
                        missingrow = rowidx+1
                    else:
                        not_missing.append(cell.value)
                #Escape/Illegal characters in data check        
                #if badchar in cell.value:
                #    has_escapes = True
                #    bad_str.append(badchar)
                #    errorcol = colidx
                #    errorrow = rowidx
                #    illegalchar = cell.value
                if cell.value == '':
                        missing = True
                        missingcol = colidx+1
                        missingrow = rowidx+1
                        missing_cells.append([missingrow, missingcol])
                else:
                    not_missing.append(cell.value)

        diff = lambda l1, l2: [x for x in l1 if x not in l2]
        missing_fields.append(str(diff(required_metadata, not_missing)))
        
        records = pe.iget_records(file_name=filename)
        # This is kinda inefficient, but we'll pre-scan the entire spreadsheet
        # before saving entries, so we don't get half-way uploaded
        # spreadsheets.
        
        #for idx, record in enumerate(records):
            # XXX: right now, we're just checking for required fields that are
            # missing, but we can add whatever checks we want here.
            # XXX: blank rows in the spreadsheet that have some hidden
            # formatting can screw up this test
            # This is where we can probably strip \r from lines and check for header accuracy
            #has_escapes = [j for j in record if '\r' in j]
            #print(record)
            
            #missing = [k for k in record if k in required_metadata and not record[k]]
            #missing = False
            #missing_fields = []
            #for i in required_metadata:
            #    if i not in record:
            #        missing = True
            #        missing_fields.append(str(i))
        
            #has_escapes = False
            #badchar = "\\"
            #bad_str = []
                    
            #for row in range(0, currentSheet.nrows):
            #    for column in "ABCDEFGHIJKLMNO":  # Here you can add or reduce the columns
            #        cell_name = "{}{}".format(column, row)
            #        if fn[cell_name].value == "\\":
                        #print("{1} cell is located on {0}" .format(cell_name, currentSheet[cell_name].value))
                        #print("cell position {} has escape character {}".format(cell_name, currentSheet[cell_name].value))
                        #return cell_name

            #for r, i in record.items():
            #    result = i
            #    if badchar in result:
            #        has_escapes = True
            #        bad_str.append(badchar)
        if missing:
            error = True
            if missing_fields[0] == '[]':
                for badcells in missing_cells:
                    error_msg = 'Missing Required Information or Extra Field found in spreadsheet in row,column "{}"'.format(badcells)
                    messages.error(request, error_msg)
            else:
                missing_str = ", ".join(missing_fields)
                error_msg = 'Data missing from row "{}" column "{}". Missing required field(s) in spreadsheet: "{}". Be sure all headers in the metadata spreadsheet provided are included and correctly spelled in your spreadsheet. If issue persists please contact us at bil-support@psc.edu.'.format(missingrow, missingcol, missing_str)
                messages.error(request, error_msg)
        if has_escapes:
            error = True
            bad_str = ", ".join(bad_str)
            error_msg = 'Data contains an illegal character in string "{}"  row: "{}" column: "{}" Be sure there are no escape characters such as "\" or "^" in your spreadsheet. If issue persists please contact us at bil-support@psc.edu.'.format(illegalchar, errorrow, errorcol)
            messages.error(request, error_msg)
        if badgrantnum:
            error = True
            error_msg = 'Grant number does not match correct format for NIH grant number, "{}" in Row: {} Column: {}  must match the format "A-B1C-2D3E4F5G-6H"'.format(grantnum, grantrow, grantcol)
            messages.error(request, error_msg)
        if error:
            # We have to add 2 to idx because spreadsheet rows are 1-indexed
            # and first row is header
            # return redirect('ingest:image_metadata_upload')
            return error
        records = pe.iget_records(file_name=filename)
        for idx, record in enumerate(records):
            im = DescriptiveMetadata(
                collection=associated_collection,
                user=request.user)
            for k in record:
                setattr(im, k, record[k])
                #messages.success(request, k)
                #messages.success(request, record[k])
            im.save()
        messages.success(request, 'Descriptive Metadata successfully uploaded')
        # return redirect('ingest:image_metadata_list')
        return error
    except pe.exceptions.FileTypeNotSupported:
        error = True
        messages.error(request, "File type not supported")
        # return redirect('ingest:image_metadata_upload')
        return error

# This gets called in the descriptive_metadata_upload function but we've commented that out to use upload_all_metadata_sheets instead, but prob will harvest some code from here. don't remove yet.
def upload_spreadsheet(spreadsheet_file, associated_collection, request):
    """ Helper used by metadata_upload and collection_detail."""
    fs = FileSystemStorage()
    filename = fs.save(spreadsheet_file.name, spreadsheet_file)
    error = False
    try:
        records = pe.iget_records(file_name=filename)
        # This is kinda inefficient, but we'll pre-scan the entire spreadsheet
        # before saving entries, so we don't get half-way uploaded
        # spreadsheets.
        for idx, record in enumerate(records):
            # XXX: right now, we're just checking for required fields that are
            # missing, but we can add whatever checks we want here.
            # XXX: blank rows in the spreadsheet that have some hidden
            # formatting can screw up this test
            missing = [k for k in record if k in required_metadata and not record[k]]
            if missing:
                error = True
                missing_str = ", ".join(missing)
                error_msg = 'Data missing from row {} in field(s): "{}"'.format(idx+2, missing_str)
                messages.error(request, error_msg)
        if error:
            # We have to add 2 to idx because spreadsheet rows are 1-indexed
            # and first row is header
            return error
        records = pe.iget_records(file_name=filename)
        for idx, record in enumerate(records):
            # "age" isn't required, so we need to explicitly set blank
            # entries to None or else django will get confused.
            if record['age'] == '':
                record['age'] = None
            im = ImageMetadata(
                collection=associated_collection,
                user=request.user)
            for k in record:
                setattr(im, k, record[k])
            im.save()
        messages.success(request, 'Metadata successfully uploaded')
        # return redirect('ingest:image_metadata_list')
        return error
    except pe.exceptions.FileTypeNotSupported:
        error = True
        messages.error(request, "File type not supported")
        return error

# DEPRECATED
# @login_required
# def image_metadata_upload(request):
#     """ Upload a spreadsheet containing image metadata information. """

#     # The POST. Auser has selected a file and associated collection to upload.
#     if request.method == 'POST' and request.FILES['spreadsheet_file']:
#         form = UploadForm(request.POST)
#         if form.is_valid():
#             collection = form.cleaned_data['associated_collection']
#             spreadsheet_file = request.FILES['spreadsheet_file']
#             error = upload_spreadsheet(spreadsheet_file, collection, project, request)
#             if error:
#                 return redirect('ingest:image_metadata_upload')
#             else:
#                 return redirect('ingest:image_metadata_list')
#     # This is the GET (just show the metadata upload page)
#     else:
#         form = UploadForm()
#         # Only let a user associate metadata with an unlocked collection that
#         # they own
#         form.fields['associated_collection'].queryset = Collection.objects.filter(
#             locked=False, user=request.user)
#     collections = Collection.objects.filter(locked=False, user=request.user)
#     return render(
#         request,
#         'ingest/image_metadata_upload.html',
#         {'form': form, 'collections': collections})

# DEPRECATED
# @login_required
# def image_metadata_list(request):
#     """ A list of all the metadata the user has created. """
#     # The user is trying to delete the selected metadata
#     if request.method == "POST":
#         pks = request.POST.getlist("selection")
#         # Get all of the checked metadata (except LOCKED metadata)
#         selected_objects = ImageMetadata.objects.filter(
#             pk__in=pks, locked=False)
#         selected_objects.delete()
#         messages.success(request, 'Metadata successfully deleted')
#         return redirect('ingest:image_metadata_list')
#     # This is the GET (just show the user their list of metadata)
#     else:
#         # XXX: This exclude is likely redundant, becaue there's already the
#         # same exclude in the class itself. Need to test though.
#         table = ImageMetadataTable(
#             ImageMetadata.objects.filter(user=request.user), exclude=['user','bil_uuid'])
#         RequestConfig(request).configure(table)
#         image_metadata = ImageMetadata.objects.filter(user=request.user)
#         return render(
#             request,
#             'ingest/image_metadata_list.html',
#             {'table': table, 'image_metadata': image_metadata})

# DEPRECATED
# class ImageMetadataDetail(LoginRequiredMixin, DetailView):
#     """ A detailed view of a single piece of metadata. """
#     model = ImageMetadata
#     template_name = 'ingest/image_metadata_detail.html'
#     context_object_name = 'image_metadata'

# DEPRECATED
# @login_required
# def image_metadata_create(request):
#     """ Create new image metadata. """
#     # The user has hit the "Save" button on the "Create Metadata" page.
#     if request.method == "POST":
#         # We need to pass in request here, so we can use it to get the user
#         form = ImageMetadataForm(request.POST, user=request.user)
#         if form.is_valid():
#             post = form.save(commit=False)
#             post.save()
#             messages.success(request, 'Metadata successfully created')
#             return redirect('ingest:image_metadata_list')
#     # The GET. Just show the user the blank "Create Metadata" form.
#     else:
#         form = ImageMetadataForm(user=request.user)
#         # Only let a user associate metadata with an unlocked collection that
#         # they own
#         form.fields['collection'].queryset = Collection.objects.filter(
#             locked=False, user=request.user)
#     return render(request, 'ingest/image_metadata_create.html', {'form': form})
# DEPRECATED
# class ImageMetadataUpdate(LoginRequiredMixin, UpdateView):
#     """ Modify an existing piece of image metadata. """
#     model = ImageMetadata
#     template_name = 'ingest/image_metadata_update.html'
#     success_url = reverse_lazy('ingest:image_metadata_list')
#     form_class = ImageMetadataForm

#     def get_form_kwargs(self):
#         kwargs = super(ImageMetadataUpdate, self).get_form_kwargs()
#         kwargs.update({'user': self.request.user})
#         return kwargs

# DEPRECATED
# class ImageMetadataDelete(LoginRequiredMixin, DeleteView):
#     """ Delete an existing piece of image metadata. """
#     model = ImageMetadata
#     template_name = 'ingest/image_metadata_delete.html'
#     success_url = reverse_lazy('ingest:image_metadata_list')
